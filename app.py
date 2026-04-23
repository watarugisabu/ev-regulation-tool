# =============================================================================
# EV充電器設置工事 規制区域自動判定ツール
# =============================================================================
# 使い方:
#   1. Python 3.10以上をインストール
#   2. ターミナルで以下を実行:
#      pip install streamlit pandas openpyxl geopandas shapely requests folium streamlit-folium
#   3. 国土数値情報から自然公園地域データ・景観計画区域データをダウンロードし
#      data/ フォルダに配置（詳細はアプリ内の説明を参照）
#   4. ターミナルで以下を実行:
#      streamlit run app.py
# =============================================================================

import streamlit as st
import pandas as pd
import requests
import time
import io
import os
import json
import glob
from datetime import datetime

# --- GISライブラリ（インストールされていない場合のフォールバック） ---
try:
    import geopandas as gpd
    from shapely.geometry import Point
    HAS_GIS = True
except ImportError:
    HAS_GIS = False

try:
    import folium
    from streamlit_folium import st_folium
    HAS_MAP = True
except ImportError:
    HAS_MAP = False


# =============================================================================
# ページ設定
# =============================================================================
st.set_page_config(
    page_title="EV設置 規制区域判定ツール",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded",
)

# =============================================================================
# カスタムCSS
# =============================================================================
st.markdown("""
<style>
    /* ヘッダー */
    .main-header {
        font-size: 1.8rem;
        font-weight: 700;
        color: #1a365d;
        padding-bottom: 0.5rem;
        border-bottom: 3px solid #3182ce;
        margin-bottom: 1rem;
    }
    /* ステータスバッジ */
    .badge-required {
        background-color: #fed7d7;
        color: #c53030;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.85rem;
        font-weight: 600;
    }
    .badge-ok {
        background-color: #c6f6d5;
        color: #276749;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.85rem;
        font-weight: 600;
    }
    .badge-unknown {
        background-color: #fefcbf;
        color: #975a16;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 0.85rem;
        font-weight: 600;
    }
    /* 統計カード */
    .stat-card {
        background: #f7fafc;
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        padding: 1rem;
        text-align: center;
    }
    .stat-number {
        font-size: 2rem;
        font-weight: 700;
        color: #2d3748;
    }
    .stat-label {
        font-size: 0.85rem;
        color: #718096;
    }
</style>
""", unsafe_allow_html=True)


# =============================================================================
# ジオコーディング（国土地理院API）
# =============================================================================
def geocode_address(address: str) -> dict:
    """
    国土地理院APIを使って住所を緯度経度に変換する。
    無料・登録不要で使えるAPI。
    """
    url = "https://msearch.gsi.go.jp/address-search/AddressSearch"
    params = {"q": address}
    try:
        res = requests.get(url, params=params, timeout=10)
        data = res.json()
        if data and len(data) > 0:
            coords = data[0]["geometry"]["coordinates"]
            title = data[0]["properties"].get("title", "")
            return {
                "success": True,
                "lat": coords[1],   # 緯度
                "lon": coords[0],   # 経度
                "matched": title,
            }
    except Exception as e:
        return {"success": False, "error": str(e)}
    return {"success": False, "error": "住所が見つかりませんでした"}


# =============================================================================
# GISデータ読み込み
# =============================================================================
@st.cache_data(ttl=3600)
def load_gis_data(data_dir: str, pattern: str, label: str):
    """
    data/ フォルダ内のShapefile/GeoJSONを読み込む。
    複数ファイルがあれば結合する。
    """
    if not HAS_GIS:
        return None

    # アプリのディレクトリ基準でもdata/を探す
    possible_dirs = [data_dir]
    app_dir = os.path.dirname(os.path.abspath(__file__))
    possible_dirs.append(os.path.join(app_dir, "data"))
    possible_dirs.append(os.path.join(app_dir, data_dir))

    files = []
    for d in possible_dirs:
        if not os.path.exists(d):
            continue
        for ext in ["*.shp", "*.geojson", "*.json"]:
            # サブフォルダも含めて検索
            found = glob.glob(os.path.join(d, "**", ext), recursive=True)
            found += glob.glob(os.path.join(d, ext))
            for f in found:
                fname = os.path.basename(f).lower()
                if label.lower() in fname and f not in files:
                    files.append(f)

    if not files:
        return None

    gdfs = []
    for f in files:
        try:
            gdf = gpd.read_file(f, encoding="utf-8")
            if gdf.crs and gdf.crs.to_epsg() != 4326:
                gdf = gdf.to_crs(epsg=4326)
            gdfs.append(gdf)
        except Exception:
            try:
                gdf = gpd.read_file(f, encoding="cp932")
                if gdf.crs and gdf.crs.to_epsg() != 4326:
                    gdf = gdf.to_crs(epsg=4326)
                gdfs.append(gdf)
            except Exception:
                pass

    if gdfs:
        return pd.concat(gdfs, ignore_index=True)
    return None


def check_point_in_areas(lat: float, lon: float, gdf, name_cols: list) -> list:
    """
    指定した緯度経度がGISポリゴン内に含まれるか判定する。
    該当するポリゴンの属性情報をリストで返す。
    """
    if gdf is None or not HAS_GIS:
        return []

    point = Point(lon, lat)
    results = []

    # ポリゴンとの照合
    mask = gdf.geometry.contains(point)
    hits = gdf[mask]

    for _, row in hits.iterrows():
        info = {}
        for col in name_cols:
            if col in row.index:
                val = row[col]
                if pd.notna(val):
                    info[col] = str(val)
        results.append(info)

    return results


# =============================================================================
# Excel出力
# =============================================================================
def create_output_excel(df: pd.DataFrame) -> bytes:
    """
    判定結果をExcelファイルとして生成する。
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 全件シート
        df.to_excel(writer, sheet_name="判定結果一覧", index=False)

        # 要申請案件シート
        required = df[df["総合判定"] == "要確認あり"]
        if len(required) > 0:
            required.to_excel(writer, sheet_name="要確認案件", index=False)

        # 集計シート
        summary_data = {
            "項目": ["総件数", "ジオコーディング成功", "ジオコーディング失敗",
                     "自然公園区域 該当", "景観計画区域 該当", "要確認案件"],
            "件数": [
                len(df),
                len(df[df["緯度"].notna()]),
                len(df[df["緯度"].isna()]),
                len(df[df["自然公園"] != "該当なし"]) if "自然公園" in df.columns else 0,
                len(df[df["景観計画区域"] != "該当なし"]) if "景観計画区域" in df.columns else 0,
                len(required),
            ]
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="集計", index=False)

        # 書式設定
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            # 列幅の自動調整（簡易版）
            for col_cells in ws.columns:
                max_length = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                adjusted = min(max_length + 4, 50)
                ws.column_dimensions[col_letter].width = adjusted
            # ヘッダー行の書式
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            # フィルタ設定
            ws.auto_filter.ref = ws.dimensions

    return output.getvalue()


# =============================================================================
# メイン画面
# =============================================================================

# --- サイドバー ---
with st.sidebar:
    st.markdown("### ⚡ EV設置 規制区域判定")
    st.markdown("---")
    st.markdown("#### 📁 GISデータの配置")

    data_dir = st.text_input(
        "データフォルダのパス",
        value="data",
        help="国土数値情報からダウンロードしたデータを置くフォルダ"
    )

    st.markdown("""
    **データのダウンロード方法:**

    1. [国土数値情報](https://nlftp.mlit.go.jp/ksj/) にアクセス
    2. 「自然公園地域」データをダウンロード
    3. 「景観計画区域」データをダウンロード
    4. 解凍して上記フォルダに配置

    ※ Shapefile(.shp)またはGeoJSON形式
    """)

    st.markdown("---")

    # GISデータの読み込み状況
    st.markdown("#### 📊 データ読み込み状況")

    if HAS_GIS and os.path.exists(data_dir):
        # 自然公園データ
        park_gdf = load_gis_data(data_dir, "*", "A10")
        if park_gdf is None:
            park_gdf = load_gis_data(data_dir, "*", "park")
        if park_gdf is None:
            park_gdf = load_gis_data(data_dir, "*", "自然公園")

        if park_gdf is not None:
            st.success(f"自然公園: {len(park_gdf)}件のポリゴン")
        else:
            st.warning("自然公園: データ未配置")

        # 景観計画区域データ
        landscape_gdf = load_gis_data(data_dir, "*", "A35")
        if landscape_gdf is None:
            landscape_gdf = load_gis_data(data_dir, "*", "A34")
        if landscape_gdf is None:
            landscape_gdf = load_gis_data(data_dir, "*", "景観")
        if landscape_gdf is None:
            landscape_gdf = load_gis_data(data_dir, "*", "landscape")

        if landscape_gdf is not None:
            st.success(f"景観計画区域: {len(landscape_gdf)}件のポリゴン")
        else:
            st.warning("景観計画区域: データ未配置")
    else:
        park_gdf = None
        landscape_gdf = None
        if not HAS_GIS:
            st.error("geopandasが未インストール")
        elif not os.path.exists(data_dir):
            st.warning(f"'{data_dir}' フォルダが見つかりません")

    st.markdown("---")
    st.markdown("#### ⚙️ 設定")
    geocode_delay = st.slider(
        "API呼び出し間隔（秒）",
        min_value=0.3,
        max_value=3.0,
        value=1.0,
        step=0.1,
        help="国土地理院APIへの負荷軽減のため間隔を空けます"
    )


# --- メインコンテンツ ---
st.markdown('<div class="main-header">⚡ EV充電器設置工事 規制区域自動判定ツール</div>', unsafe_allow_html=True)

st.markdown("""
住所リスト（Excel/CSV）をアップロードすると、各住所が以下の規制区域に該当するかを自動判定します。
- **自然公園区域**（国立公園・国定公園・都道府県立自然公園）
- **景観計画区域**（環境色対応が必要な地域）
""")

# --- ファイルアップロード ---
st.markdown("### 📂 住所リストのアップロード")

uploaded_file = st.file_uploader(
    "Excel(.xlsx)またはCSV(.csv)ファイルを選択してください",
    type=["xlsx", "csv"],
    help="「住所」列を含むファイルをアップロードしてください"
)

if uploaded_file is not None:
    # ファイル読み込み
    try:
        if uploaded_file.name.endswith(".csv"):
            # CSVはエンコーディングを自動判定
            try:
                df = pd.read_csv(uploaded_file, encoding="utf-8")
            except UnicodeDecodeError:
                uploaded_file.seek(0)
                df = pd.read_csv(uploaded_file, encoding="cp932")
        else:
            df = pd.read_excel(uploaded_file)

        st.success(f"ファイルを読み込みました: {len(df)}行 × {len(df.columns)}列")

        # 住所列の特定
        st.markdown("### 📍 住所列の選択")
        address_col_candidates = [
            col for col in df.columns
            if any(kw in str(col) for kw in ["住所", "アドレス", "address", "所在地", "場所", "Address"])
        ]

        if address_col_candidates:
            default_idx = 0
        else:
            address_col_candidates = list(df.columns)
            default_idx = 0

        address_col = st.selectbox(
            "住所が入っている列を選択してください",
            options=df.columns.tolist(),
            index=df.columns.tolist().index(address_col_candidates[0]) if address_col_candidates else 0,
        )

        # プレビュー
        with st.expander("📋 データプレビュー（先頭10行）", expanded=True):
            st.dataframe(df.head(10), use_container_width=True)

        # --- 判定実行 ---
        st.markdown("### 🔍 判定の実行")

        col1, col2 = st.columns([1, 3])
        with col1:
            run_button = st.button(
                "▶ 判定を開始",
                type="primary",
                use_container_width=True,
            )

        if run_button:
            results = []
            total = len(df)

            progress_bar = st.progress(0, text="判定を実行中...")
            status_text = st.empty()

            for idx, row in df.iterrows():
                address = str(row[address_col]).strip()

                if not address or address == "nan":
                    results.append({
                        "元の住所": address,
                        "マッチ住所": "",
                        "緯度": None,
                        "経度": None,
                        "ジオコーディング": "住所なし",
                        "自然公園": "判定不可",
                        "自然公園_詳細": "",
                        "景観計画区域": "判定不可",
                        "景観計画区域_詳細": "",
                        "総合判定": "判定不可",
                    })
                    continue

                status_text.text(f"処理中: {idx + 1}/{total} - {address}")
                progress_bar.progress((idx + 1) / total)

                # ジオコーディング
                geo = geocode_address(address)

                if geo["success"]:
                    lat, lon = geo["lat"], geo["lon"]

                    # 自然公園判定
                    park_result = "該当なし"
                    park_detail = ""
                    if park_gdf is not None:
                        # 自然公園データの列名候補
                        park_name_cols = [
                            c for c in park_gdf.columns
                            if any(kw in str(c).lower() for kw in
                                   ["nam", "名", "name", "公園", "park",
                                    "種別", "type", "区域", "zone", "地域"])
                        ]
                        if not park_name_cols:
                            park_name_cols = [c for c in park_gdf.columns if c != "geometry"]

                        hits = check_point_in_areas(lat, lon, park_gdf, park_name_cols)
                        if hits:
                            park_result = "該当あり"
                            details = []
                            for h in hits:
                                details.append(" / ".join(f"{k}:{v}" for k, v in h.items()))
                            park_detail = " | ".join(details)

                    # 景観計画区域判定
                    landscape_result = "該当なし"
                    landscape_detail = ""
                    if landscape_gdf is not None:
                        landscape_name_cols = [
                            c for c in landscape_gdf.columns
                            if any(kw in str(c).lower() for kw in
                                   ["nam", "名", "name", "景観", "land",
                                    "区域", "zone", "自治体", "市", "区"])
                        ]
                        if not landscape_name_cols:
                            landscape_name_cols = [c for c in landscape_gdf.columns if c != "geometry"]

                        hits = check_point_in_areas(lat, lon, landscape_gdf, landscape_name_cols)
                        if hits:
                            landscape_result = "該当あり"
                            details = []
                            for h in hits:
                                details.append(" / ".join(f"{k}:{v}" for k, v in h.items()))
                            landscape_detail = " | ".join(details)

                    # データ未配置の場合の表示
                    if park_gdf is None:
                        park_result = "データ未配置"
                    if landscape_gdf is None:
                        landscape_result = "データ未配置"

                    # 総合判定
                    has_issue = (park_result == "該当あり" or landscape_result == "該当あり")
                    overall = "要確認あり" if has_issue else "申請不要の可能性"
                    if park_gdf is None and landscape_gdf is None:
                        overall = "GISデータ未配置のため判定不可"

                    results.append({
                        "元の住所": address,
                        "マッチ住所": geo.get("matched", ""),
                        "緯度": lat,
                        "経度": lon,
                        "ジオコーディング": "成功",
                        "自然公園": park_result,
                        "自然公園_詳細": park_detail,
                        "景観計画区域": landscape_result,
                        "景観計画区域_詳細": landscape_detail,
                        "総合判定": overall,
                    })
                else:
                    results.append({
                        "元の住所": address,
                        "マッチ住所": "",
                        "緯度": None,
                        "経度": None,
                        "ジオコーディング": f"失敗: {geo.get('error', '')}",
                        "自然公園": "判定不可",
                        "自然公園_詳細": "",
                        "景観計画区域": "判定不可",
                        "景観計画区域_詳細": "",
                        "総合判定": "判定不可",
                    })

                # API負荷軽減のため待機
                time.sleep(geocode_delay)

            progress_bar.progress(1.0, text="判定完了！")
            status_text.text("")

            # 結果をDataFrameに変換
            result_df = pd.DataFrame(results)
            st.session_state["result_df"] = result_df
            st.session_state["original_df"] = df

        # --- 結果表示 ---
        if "result_df" in st.session_state:
            result_df = st.session_state["result_df"]

            st.markdown("---")
            st.markdown("### 📊 判定結果")

            # 統計サマリー
            total = len(result_df)
            geo_ok = len(result_df[result_df["ジオコーディング"] == "成功"])
            geo_fail = total - geo_ok
            park_hit = len(result_df[result_df["自然公園"] == "該当あり"])
            landscape_hit = len(result_df[result_df["景観計画区域"] == "該当あり"])
            need_check = len(result_df[result_df["総合判定"] == "要確認あり"])

            cols = st.columns(5)
            with cols[0]:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{total}</div>
                    <div class="stat-label">総件数</div>
                </div>""", unsafe_allow_html=True)
            with cols[1]:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number">{geo_ok}</div>
                    <div class="stat-label">住所変換成功</div>
                </div>""", unsafe_allow_html=True)
            with cols[2]:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:#c53030">{park_hit}</div>
                    <div class="stat-label">自然公園 該当</div>
                </div>""", unsafe_allow_html=True)
            with cols[3]:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:#c53030">{landscape_hit}</div>
                    <div class="stat-label">景観計画 該当</div>
                </div>""", unsafe_allow_html=True)
            with cols[4]:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:#c53030">{need_check}</div>
                    <div class="stat-label">要確認案件</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("")

            # フィルター
            filter_option = st.radio(
                "表示フィルター",
                ["全件表示", "要確認案件のみ", "ジオコーディング失敗のみ"],
                horizontal=True,
            )

            if filter_option == "要確認案件のみ":
                display_df = result_df[result_df["総合判定"] == "要確認あり"]
            elif filter_option == "ジオコーディング失敗のみ":
                display_df = result_df[result_df["ジオコーディング"] != "成功"]
            else:
                display_df = result_df

            # 結果テーブル（色付き表示）
            def highlight_row(row):
                if row["総合判定"] == "要確認あり":
                    return ["background-color: #fff5f5"] * len(row)
                elif row["総合判定"] == "判定不可":
                    return ["background-color: #fffff0"] * len(row)
                else:
                    return ["background-color: #f0fff4"] * len(row)

            # 表示用の列を選択（詳細列は折りたたむ）
            show_cols = ["元の住所", "マッチ住所", "緯度", "経度",
                         "自然公園", "景観計画区域", "総合判定"]
            display_styled = display_df[show_cols].style.apply(highlight_row, axis=1)
            st.dataframe(display_styled, use_container_width=True, height=400)

            # 詳細情報（展開可能）
            with st.expander("📝 詳細情報（自然公園名・景観区域名など）"):
                detail_cols = ["元の住所", "自然公園", "自然公園_詳細",
                               "景観計画区域", "景観計画区域_詳細"]
                detail_df = result_df[result_df["自然公園_詳細"].str.len() > 0 |
                                      (result_df["景観計画区域_詳細"].str.len() > 0)]
                if len(detail_df) > 0:
                    st.dataframe(detail_df[detail_cols], use_container_width=True)
                else:
                    st.info("該当する詳細情報はありません")

            # --- 地図表示 ---
            st.markdown("### 🗺️ 地図表示")

            map_df = result_df[result_df["緯度"].notna()].copy()

            if len(map_df) > 0:
                if HAS_MAP:
                    # 中心座標を計算
                    center_lat = map_df["緯度"].mean()
                    center_lon = map_df["経度"].mean()

                    m = folium.Map(
                        location=[center_lat, center_lon],
                        zoom_start=6,
                        tiles="https://cyberjapandata.gsi.go.jp/xyz/pale/{z}/{x}/{y}.png",
                        attr="国土地理院",
                    )

                    for _, row in map_df.iterrows():
                        # マーカーの色を判定結果で変える
                        if row["総合判定"] == "要確認あり":
                            color = "red"
                            icon = "exclamation-sign"
                        elif row["総合判定"] == "判定不可":
                            color = "orange"
                            icon = "question-sign"
                        else:
                            color = "green"
                            icon = "ok-sign"

                        popup_html = f"""
                        <div style="width:250px; font-size:12px;">
                            <b>{row['元の住所']}</b><br>
                            <hr style="margin:4px 0">
                            自然公園: {row['自然公園']}<br>
                            景観計画: {row['景観計画区域']}<br>
                            <b>判定: {row['総合判定']}</b>
                        </div>
                        """

                        folium.Marker(
                            location=[row["緯度"], row["経度"]],
                            popup=folium.Popup(popup_html, max_width=300),
                            tooltip=row["元の住所"][:30],
                            icon=folium.Icon(color=color, icon=icon),
                        ).add_to(m)

                    st_folium(m, width=None, height=500, use_container_width=True)

                    st.markdown("""
                    🔴 **赤** = 要確認（規制区域に該当）&nbsp;&nbsp;
                    🟢 **緑** = 申請不要の可能性&nbsp;&nbsp;
                    🟠 **橙** = 判定不可
                    """)
                else:
                    st.info("地図表示にはfoliumとstreamlit-foliumのインストールが必要です")
                    st.code("pip install folium streamlit-folium", language="bash")
            else:
                st.warning("地図表示できる住所がありません（ジオコーディング結果を確認してください）")

            # --- Excel出力 ---
            st.markdown("### 📥 結果のダウンロード")

            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"判定結果_{now}.xlsx"

            excel_data = create_output_excel(result_df)

            st.download_button(
                label="📥 判定結果をExcelでダウンロード",
                data=excel_data,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
                type="primary",
                use_container_width=True,
            )

            st.markdown("""
            **Excelファイルの内容:**
            - 「判定結果一覧」シート — 全住所の判定結果
            - 「要確認案件」シート — 規制区域に該当する案件のみ
            - 「集計」シート — 規制種別ごとの件数サマリー
            """)

    except Exception as e:
        st.error(f"ファイルの読み込みに失敗しました: {str(e)}")

else:
    # ファイル未アップロード時の案内
    st.markdown("---")
    st.info("👆 まずExcelまたはCSVファイルをアップロードしてください")

    st.markdown("### 📝 入力ファイルの形式")
    st.markdown("住所列を含むExcelまたはCSVファイルを用意してください。例:")

    sample_data = pd.DataFrame({
        "No.": [1, 2, 3, 4, 5],
        "施設名": [
            "○○ショッピングモール",
            "△△コンビニ □□店",
            "◇◇ホテル",
            "××スーパー",
            "☆☆道の駅",
        ],
        "住所": [
            "東京都新宿区西新宿二丁目8番1号",
            "神奈川県箱根町湯本682",
            "栃木県日光市中宮祠2482",
            "大阪府大阪市中央区難波1-1-1",
            "長野県松本市安曇4306",
        ],
    })
    st.dataframe(sample_data, use_container_width=True)

    st.markdown("""
    **ポイント:**
    - 「住所」という列名が含まれていると自動で検出されます
    - 列名が異なる場合は、アップロード後に選択できます
    - 都道府県名から始まる住所が最も変換精度が高くなります
    """)

# --- フッター ---
st.markdown("---")
st.markdown(
    '<p style="text-align:center; color:#a0aec0; font-size:0.8rem;">'
    'EV充電器設置工事 規制区域自動判定ツール v1.0 | '
    'GISデータ出典: 国土数値情報（国土交通省）| '
    'ジオコーディング: 国土地理院'
    '</p>',
    unsafe_allow_html=True
)
